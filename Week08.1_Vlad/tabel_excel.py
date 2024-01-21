import openpyxl as xl
from openpyxl.chart import BarChart, Reference


numere = [chr(n) for n in range(ord('0'), ord('0')+10)]
litere = [chr(n) for n in range(ord('a'), ord('a')+26)]
semne = [chr(n) for n in range(ord("!"), ord("!")+15)] + [chr(n) for n in range(ord(":"), ord(":")+7)] + [chr(n)
                                                                                                          for n in range(ord("["), ord("[")+6)] + [chr(n) for n in range(ord("{"), ord("{")+4)]


colectie = {
    "numere": {},
    "litere": {},
    "semne": {},
    "alte_caractere": {}
}


with open("Week8.1_Vlad/wiatrace.log") as f:
    for linie in f.readlines():
        linie.strip()
        for caracter in linie:
            if caracter.lower() in litere:
                if caracter.lower() in colectie["litere"].keys():
                    colectie["litere"][f"{caracter.lower()}"] += 1
                else:
                    colectie["litere"][f"{caracter.lower()}"] = 1
            elif caracter in numere:
                if caracter in colectie["numere"].keys():
                    colectie["numere"][f"{caracter}"] += 1
                else:
                    colectie["numere"][f"{caracter}"] = 1
            elif caracter in semne:
                if caracter in colectie["semne"].keys():
                    colectie["semne"][f"{caracter}"] += 1
                else:
                    colectie["semne"][f"{caracter}"] = 1
            else:
                if caracter in colectie["alte_caractere"].keys():
                    colectie["alte_caractere"][f"{caracter}"] += 1
                else:
                    colectie["alte_caractere"][f"{caracter}"] = 1

# sortez elementele
for elem in colectie.keys():
    colectie[elem] = dict(sorted(colectie[elem].items()))

workbook = xl.load_workbook("Week8.1_Vlad/fisier.xlsx")
sheet = workbook["Sheet1"]

# asta e un obiect si nu reprezinta valoarea din celula
cell = sheet.cell(1, 1)
cell.value = "Litera"
cell = sheet.cell(1, 2)
cell.value = "Frecventa"

cell = sheet.cell(1, 15)
cell.value = "Cifra"
cell = sheet.cell(1, 16)
cell.value = "Frecventa"

cell = sheet.cell(1, 28)
cell.value = "Semnul"
cell = sheet.cell(1, 29)
cell.value = "Frecventa"

# fac o lista de tuple (key, value) ca sa pot itera peste valori
lista_valori_litere = []
lista_valori_numere = []
lista_valori_semne = []

# fac aici o tupla ca sa pot sa iterez prin elemete
for key, value in colectie["litere"].items():
    lista_valori_litere.append((key, value))

for key, value in colectie["numere"].items():
    lista_valori_numere.append((key, value))

for key, value in colectie["semne"].items():
    lista_valori_semne.append((key, value))


for index in range(len(lista_valori_litere)):
    col1_cell = sheet.cell(index+2, 1)
    col2_cell = sheet.cell(index+2, 2)
    col1_cell.value = lista_valori_litere[index][0]
    col2_cell.value = lista_valori_litere[index][1]

for index in range(len(lista_valori_numere)):
    col1_cell = sheet.cell(index+2, 15)
    col2_cell = sheet.cell(index+2, 16)
    col1_cell.value = lista_valori_numere[index][0]
    col2_cell.value = lista_valori_numere[index][1]

for index in range(len(lista_valori_semne)):
    col1_cell = sheet.cell(index+2, 28)
    col2_cell = sheet.cell(index+2, 29)
    col1_cell.value = lista_valori_semne[index][0]
    col2_cell.value = lista_valori_semne[index][1]

values_litere = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=2, max_col=2)
categorii_litere = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)

values_numere = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=16, max_col=16)
categorii_numere = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=15, max_col=15)

values_semne = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=29, max_col=29)
categorii_semne = Reference(
    sheet, min_row=2, max_row=sheet.max_row, min_col=28, max_col=28)


chart_litere = BarChart()
chart_litere.title = "Frecventa literelor"
chart_litere.y_axis.title = "Frecventa"
chart_litere.x_axis.title = "Litera"
chart_litere.add_data(values_litere)
chart_litere.set_categories(categorii_litere)
sheet.add_chart(chart_litere, "d2")

chart_numere = BarChart()
chart_numere.title = "Frecventa cifra"
chart_numere.y_axis.title = "Frecventa"
chart_numere.x_axis.title = "Cifra"
chart_numere.add_data(values_numere)
chart_numere.set_categories(categorii_numere)
sheet.add_chart(chart_numere, "r2")

chart_semne = BarChart()
chart_semne.title = "Frecventa semnelor"
chart_semne.y_axis.title = "Frecventa"
chart_semne.x_axis.title = "Semnul"
chart_semne.add_data(values_semne)
chart_semne.set_categories(categorii_semne)
sheet.add_chart(chart_semne, "ae2")

workbook.save("Week8.1_Vlad/fisier.xlsx")
